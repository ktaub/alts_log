import re
import pandas as pd
import os
import sys
# Add the code directory to the Python path to ensure imports work
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from addepar_utils import fetch_addepar_data, set_up_environment, get_portfolio_data_using_jobs
import warnings
import glob
from datetime import datetime
import openpyxl

# Suppress warnings - compatible with pandas 1.x and 2.x
warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", category=UserWarning)
try:
    warnings.filterwarnings("ignore", category=pd.errors.SettingWithCopyWarning)
except AttributeError:
    # SettingWithCopyWarning doesn't exist in pandas 2.0+
    pass

TRANSACTIONS_CACHE_FILE = 'transactions_cache.xlsx'


def is_false_value(x):
    """Check if value represents False - handles 0, '0', 'false', False, 0.0"""
    if pd.isna(x):
        return False
    str_val = str(x).strip().lower()
    return str_val in ['false', '0', '0.0', 'no', 'n']


def coerce_id_columns_to_int64(df):
    """Ensure ID columns are numeric for consistent joins."""
    for col in ['Direct Owner Entity ID', 'Entity ID']:
        if col in df.columns:
            # Use float64 for better compatibility across pandas versions
            # NaN values are preserved, which is what we need for outer joins
            df[col] = pd.to_numeric(df[col], errors='coerce').astype('float64')
    return df


def get_cached_transactions(cache_file=TRANSACTIONS_CACHE_FILE):
    """Load cached transactions and return DataFrame and last Trade Date."""
    if os.path.exists(cache_file):
        try:
            if os.path.getsize(cache_file) == 0:
                return None, None
            cached_df = pd.read_excel(cache_file, engine="openpyxl")
        except Exception:
            return None, None
        return cached_df, None
    return None, None


def save_cached_transactions(df, cache_file=TRANSACTIONS_CACHE_FILE):
    """Save transactions DataFrame to cache file."""
    df.to_excel(cache_file, index=False)


def fetch_and_process_addepar_data(api_key, api_secret, firm_id, base_url, start_date, end_date):
    """Fetch and process transaction data from Addepar, using cache if available."""
    print("\n[STEP 1/7] Loading cached transaction data...")
    cached_df, _ = get_cached_transactions(TRANSACTIONS_CACHE_FILE)
    if cached_df is not None:
        print(f"  [OK] Loaded {len(cached_df)} cached transactions")
    else:
        print("  [INFO] No cache found")
    
    today_dt = pd.to_datetime(end_date)
    fetch_start_date = (today_dt - pd.Timedelta(days=5)).strftime('%Y-%m-%d')

    print(f"[STEP 1/7] Fetching new transactions from Addepar (from {fetch_start_date} to {end_date})...")
    # Only fetch if fetch_start_date <= end_date
    transactions_df, success, message = fetch_addepar_data(
        "transactions", "105390", api_key, api_secret, firm_id, base_url, fetch_start_date, end_date
    )

    if not success:
        print(f"  [ERROR] Failed to fetch transactions: {message}")
        # If cache exists, use cached data only
        if cached_df is not None:
            print("  [INFO] Using cached data only")
            transactions_df = cached_df
        else:
            return None, None
    else:
        print(f"  [OK] Fetched {len(transactions_df)} new transactions")
        # Combine cached and new data
        if cached_df is not None:
            print("[STEP 1/7] Combining cached and new transactions...")
            combined_df = pd.concat(
                [cached_df, transactions_df], ignore_index=True)
            combined_df = combined_df.sort_values(
                'Posted Date').drop_duplicates('ID', keep='last')
            transactions_df = combined_df
            print(f"  [OK] Combined total: {len(transactions_df)} transactions")
        # Save updated cache
        print("[STEP 1/7] Saving updated cache...")
        save_cached_transactions(transactions_df, TRANSACTIONS_CACHE_FILE)
        print("  [OK] Cache saved")

    # Standardize ID columns as numeric for joins
    print("[STEP 1/7] Processing transaction data...")
    transactions_df = coerce_id_columns_to_int64(transactions_df)

    # Only include Distributions where Return of Capital is not zero
    mask = ~((transactions_df["Type"] == "Distribution") & (
        transactions_df["Return of Capital"] == 0))
    transactions_df = transactions_df[mask]

    # Group by Direct Owner and Security (names, not IDs)
    tx_grouped = transactions_df.groupby(["Direct Owner Entity ID", "Entity ID", "Type"]).agg({
        "Value": "sum", "Trade Date": "max"}).reset_index()
    tx_grouped = tx_grouped.rename(columns={"Value": "Total"})

    # Create a pivot table: one row per position, columns for each transaction type
    pivot = tx_grouped.pivot_table(
        index=["Direct Owner Entity ID", "Entity ID"],
        columns="Type",
        values="Total",
        aggfunc="sum",
        fill_value=0
    ).reset_index()

    # Convert all values to absolute
    for col in pivot.columns:
        if col in tx_grouped["Type"].unique():
            pivot[col] = pivot[col].abs()

    # Subtract cancellation columns from original transaction columns
    def adjust_for_cancellation(df, txn_type):
        orig_col = txn_type
        cancel_col = f'(Cancellation) {txn_type}'
        if orig_col in df.columns:
            if cancel_col in df.columns:
                df[orig_col] = df[orig_col] - df[cancel_col]
                df.drop(columns=[cancel_col], inplace=True)
            # Ensure no negative values after subtraction
            df[orig_col] = df[orig_col].clip(lower=0)
        return df

    for txn_type in ['Buy', 'Contribution', 'Distribution', 'Sell', 'Corporate Action', 'Transfer In']:
        pivot = adjust_for_cancellation(pivot, txn_type)

    # Find last buy/contribution date
    buy_contrib = tx_grouped[tx_grouped["Type"].isin(["Buy", "Contribution"])]
    last_buy_contrib = buy_contrib.groupby(["Direct Owner Entity ID", "Entity ID"])[
        "Trade Date"].max().reset_index()
    last_buy_contrib = last_buy_contrib.rename(
        columns={"Trade Date": "Last Buy/Contribution"})

    # Find last sell/distribution date
    sell_dist = tx_grouped[tx_grouped["Type"].isin(["Sell", "Distribution"])]
    last_sell_dist = sell_dist.groupby(["Direct Owner Entity ID", "Entity ID"])[
        "Trade Date"].max().reset_index()
    last_sell_dist = last_sell_dist.rename(
        columns={"Trade Date": "Last Sell/Distribution"})
    # Check if there are any Corporate Action transactions
    corporate_action_transfer_in_in_txn = tx_grouped[tx_grouped["Type"].isin(
        ["Corporate Action", "Transfer In"])]
    has_corporate_action_transfer_in_in_txn = corporate_action_transfer_in_in_txn.groupby(["Direct Owner Entity ID", "Entity ID"])[
        "Total"].count().reset_index()
    has_corporate_action_transfer_in_in_txn = has_corporate_action_transfer_in_in_txn.rename(
        columns={"Total": "Has all transactions"})

    # Merge into the pivot table
    result = pivot.merge(last_buy_contrib, on=[
                         "Direct Owner Entity ID", "Entity ID"], how="left")
    result = result.merge(last_sell_dist, on=[
                          "Direct Owner Entity ID", "Entity ID"], how="left")
    result = result.merge(has_corporate_action_transfer_in_in_txn, on=[
                          "Direct Owner Entity ID", "Entity ID"], how="left")
    # Standardize ID columns for the result as well
    result = coerce_id_columns_to_int64(result)
    print(f"  [OK] Transaction processing complete - {len(result)} positions\n")

    return transactions_df, result


def fetch_and_process_alts_list_data(api_key, api_secret, firm_id, base_url, start_date, end_date):
    """Fetch and process transaction data from Addepar, using cache if available."""
    print("[STEP 2/7] Fetching alts list data from Addepar...")
    today_dt = pd.to_datetime(end_date)
    fetch_start_date = (today_dt - pd.Timedelta(days=7)).strftime('%Y-%m-%d')

    # Only fetch if fetch_start_date <= end_date
    success, message, alts_list_df = get_portfolio_data_using_jobs(
        "portfolio", "734752", api_key, api_secret, firm_id, base_url, fetch_start_date, end_date
    )
    
    if not success or alts_list_df is None:
        print(f"  [ERROR] Failed to fetch alts list data: {message}")
        return None
    
    print(f"  [OK] Fetched {len(alts_list_df)} alts list records")

    # remove last row
    alts_list_df = alts_list_df.iloc[:-1]

    # rename Direct Owner ID to Direct Owner Entity ID
    alts_list_df = alts_list_df.rename(columns={"Direct Owner ID": "Direct Owner Entity ID", "Capital Returned (Since Inception, USD)": "Capital Returned",
                                                "Total Commitments (Since Inception, USD)": "Total Commitments", "Total Contributions (Since Inception, USD)": "Total Contributions"})
    
    # Convert ID columns to numeric for consistent joins
    alts_list_df = coerce_id_columns_to_int64(alts_list_df)
    
    print(f"  [OK] Alts list processing complete\n")
    return alts_list_df


def process_alts_info_data():
    """Load and process the Alts Info Excel file."""
    print("[STEP 3/7] Loading and processing Alts Info Excel file...")
    # Read both sheets from Alts Info file
    alternatives_df = pd.read_excel(
        "Alts Info.xlsx", sheet_name="Alternatives", engine="openpyxl")
    
    historical_df = pd.read_excel(
        "Alts Info.xlsx", sheet_name="Historical", engine="openpyxl")
    print(f"  [OK] Loaded {len(historical_df)} historical records")

    # add rows from historical_df to alternatives_df
    alternatives_df = pd.concat([alternatives_df, historical_df])
    print(f"  [OK] Added {len(historical_df)} historical records to alternatives_df")

    # Remove empty rows immediately to improve performance
    print(f"  - Original rows: {len(alternatives_df)}")
    # Drop rows where all values are NaN or where key columns are empty
    alternatives_df = alternatives_df.dropna(how='all')
    print(f"  - After removing empty rows: {len(alternatives_df)}")

    # Combine the sheets
    investment_status_df = alternatives_df
    investment_status_df = coerce_id_columns_to_int64(investment_status_df)

    # Coerce Account, Instrument, Client/Top Level Owner, Investment, Received in Addepar to string, strip whitespace, drop rows where Account/Instrument is empty or null
    investment_status_df['Account'] = investment_status_df['Account'].astype(
        str).str.strip()
    investment_status_df['Instrument'] = investment_status_df['Instrument'].astype(
        str).str.strip()

    if 'Investment' in investment_status_df.columns:
        investment_status_df['Investment'] = investment_status_df['Investment'].astype(
            str).str.strip()
    else:
        investment_status_df['Investment'] = ''
    # Handle trade date column rename: if only First Trade Date exists, use it as Last Trade Date input
    if 'Last Trade Date' not in investment_status_df.columns and 'First Trade Date' in investment_status_df.columns:
        investment_status_df['Last Trade Date'] = investment_status_df['First Trade Date']
    if 'Received in Addepar' in investment_status_df.columns:
        investment_status_df['Received in Addepar'] = investment_status_df['Received in Addepar'].astype(
            str).str.strip()
    else:
        investment_status_df['Received in Addepar'] = ''
    # Keep rows without IDs to allow pending Buys (not yet in Addepar) to appear

    # Create a normalized instrument column for matching (use Investment when not yet in Addepar)
    instrument_match = investment_status_df['Instrument']
    mask_not_in_addepar = investment_status_df['Received in Addepar'].str.lower(
    ) == 'false'
    if 'Investment' in investment_status_df.columns:
        instrument_match = instrument_match.mask(
            mask_not_in_addepar, investment_status_df['Investment'])
    investment_status_df['Instrument_Match'] = instrument_match.astype(
        str).str.strip()

    # Convert Documentation Approval Date to datetime for calculations
    investment_status_df['Documentation Approval Date'] = pd.to_datetime(
        investment_status_df['Documentation Approval Date'], errors='coerce')

    # Normalize transaction type for robust matching
    if 'Transaction Type' in investment_status_df.columns:
        investment_status_df['Txn_Type'] = investment_status_df['Transaction Type'].astype(
            str).str.strip().str.lower()
    else:
        investment_status_df['Txn_Type'] = ''

    # Create Subscription Approval Date from Buy and New Pos transactions
    new_pos_transactions = investment_status_df[investment_status_df['Transaction Type'] == 'New Pos'].copy(
    )

    # Split into received and not received in Addepar
    received_mask = new_pos_transactions['Received in Addepar'].str.lower(
    ) != 'false'

    # Get subscription dates for positions in Addepar (using IDs)
    addepar_dates = new_pos_transactions[received_mask].groupby(
        ["Direct Owner Entity ID", "Entity ID"])['Documentation Approval Date'].first().reset_index()

    # Get subscription dates for positions not in Addepar (using Top Level Owner and Investment)
    non_addepar_dates = new_pos_transactions[~received_mask].groupby(
        ["Top Level Owner", "Investment"])['Documentation Approval Date'].first().reset_index()
    non_addepar_dates['Direct Owner Entity ID'] = pd.NA
    non_addepar_dates['Entity ID'] = pd.NA

    # Combine and rename column - filter out empty DataFrames to avoid FutureWarning
    dataframes_to_concat = []
    if not addepar_dates.empty:
        dataframes_to_concat.append(addepar_dates)
    if not non_addepar_dates.empty:
        dataframes_to_concat.append(non_addepar_dates)

    if dataframes_to_concat:
        subscription_dates = pd.concat(dataframes_to_concat, ignore_index=True)
    else:
        # Create empty DataFrame with expected columns
        subscription_dates = pd.DataFrame(columns=[
                                          'Direct Owner Entity ID', 'Entity ID', 'Top Level Owner', 'Investment', 'Subscription Approval Date'])
    subscription_dates.rename(
        columns={'Documentation Approval Date': 'Subscription Approval Date'}, inplace=True)

    # Create open/close summary: for each (Account, Instrument), is there any Fully Funded? == False?
    funded_col = 'Fully Funded?'
    if funded_col in investment_status_df.columns:
        complete_summary = investment_status_df.groupby(["Direct Owner Entity ID", "Entity ID"], dropna=False)[
            funded_col].apply(lambda x: x.apply(is_false_value).any()).reset_index()
        complete_summary = complete_summary.rename(
            columns={funded_col: 'is_open'})
    else:
        # If neither column exists, default to closed
        complete_summary = investment_status_df.groupby(
            ["Direct Owner Entity ID", "Entity ID"], dropna=False).size().reset_index(name='count')
        complete_summary['is_open'] = False
        complete_summary = complete_summary.drop(columns=['count'])
    # Ensure IDs remain numeric for downstream joins
    complete_summary = coerce_id_columns_to_int64(complete_summary)
    
    print(f"  [OK] Processed {len(investment_status_df)} investment records")
    print(f"  [OK] Found {len(subscription_dates)} subscription dates\n")


    return investment_status_df, subscription_dates, complete_summary


def consolidate_alts_info_data(investment_status_df, transactions_df):
    """Consolidate Alts Info data to one row per position with calculated remaining to sell.
    For Received in Addepar == False, group by Top Level Owner + Instrument; otherwise, by Account + Instrument."""
    consolidated_positions = []

    # Split DataFrame
    received_false = investment_status_df[investment_status_df['Received in Addepar'].str.lower(
    ) == 'false'].copy()
    received_true = investment_status_df[investment_status_df['Received in Addepar'].str.lower(
    ) != 'false'].copy()

    # For items not yet in Addepar, they may lack IDs. Keep these for Open tab if Buy incomplete.
    # Already handled by keeping rows without IDs in process_alts_info_data.

    # Use Investment as Instrument for the 'false' group
    if not received_false.empty:
        received_false['Instrument'] = received_false['Investment']

    def build_positions(df, group_keys):
        if df.empty:
            return

        # Filter out blank identifiers per grouping style
        if 'Top Level Owner' in group_keys:
            df = df[
                (df['Top Level Owner'].astype(str).str.strip() != '') &
                (df['Instrument'].astype(str).str.strip() != '')
            ]
        else:
            df = df[
                df['Direct Owner Entity ID'].notna() &
                df['Entity ID'].notna()
            ]
        if df.empty:
            return

        for _, grp in df.groupby(group_keys, dropna=False):
            grp = grp.copy()

            account = str(grp['Account'].iloc[0]
                          ) if 'Account' in grp.columns else ''
            instrument = str(grp['Instrument'].iloc[0])
            direct_owner_entity_id = grp['Direct Owner Entity ID'].iloc[
                0] if 'Direct Owner Entity ID' in grp.columns else pd.NA
            entity_id = grp['Entity ID'].iloc[0] if 'Entity ID' in grp.columns else pd.NA

            owner_series = grp['Top Level Owner'].astype(str).str.strip().replace(
                '', pd.NA).dropna() if 'Top Level Owner' in grp.columns else pd.Series([], dtype=str)
            owner_value = owner_series.iloc[0] if not owner_series.empty else ''
            
            # Get Top Level Owner ID value
            owner_id_value = ''
            if 'Top Level Owner ID' in grp.columns:
                owner_id_series = grp['Top Level Owner ID'].astype(
                    str).str.strip().replace('', pd.NA).dropna()
                owner_id_value = owner_id_series.iloc[0] if not owner_id_series.empty else ''

            complete_val = None
            if 'Completed?' in grp.columns and not grp['Completed?'].isnull().all():
                complete_val = grp['Completed?'].iloc[0]

            # Get Legal Entity value - try Client/Top Level Owner first, then Top Level Owner
            legal_entity_value = ''
            if 'Top Level Owner' in grp.columns:
                legal_entity_series = grp['Top Level Owner'].astype(
                    str).str.strip().replace('', pd.NA).dropna()
                legal_entity_value = legal_entity_series.iloc[
                    0] if not legal_entity_series.empty else owner_value
            else:
                legal_entity_value = owner_value
            
            # Get Advisor value
            advisor_value = ''
            if 'Advisor' in grp.columns:
                advisor_series = grp['Advisor'].astype(
                    str).str.strip().replace('', pd.NA).dropna()
                advisor_value = advisor_series.iloc[0] if not advisor_series.empty else ''

            row = {
                'Direct Owner Entity ID': direct_owner_entity_id,
                'Entity ID': entity_id,
                'Account': account,
                'Top Level Owner': owner_value,
                'Top Level Owner ID': owner_id_value,
                'Legal Entity': legal_entity_value,
                'Advisor': advisor_value,
                'Instrument': instrument,
                'Original Commitment': 0,
                'Instruction Date': None,
                'Documentation Approval Date': None,
                'Last Trade Date': None,
                'Completed?': complete_val
            }

            # Buys and New Positions (both are buy transactions)
            buy = grp[grp['Transaction Type'].isin(['Buy', 'New Pos'])]
            if not buy.empty:
                buy = buy.sort_values(
                    'Documentation Approval Date', ascending=False)
                latest_buy = buy.iloc[0]
                row['Original Commitment'] = pd.to_numeric(
                    buy['Original Commitment'], errors='coerce').fillna(0).sum()
                row['Instruction Date'] = latest_buy.get('Instruction Date')
                row['Documentation Approval Date'] = latest_buy.get(
                    'Documentation Approval Date')
                row['Last Trade Date'] = latest_buy.get('Last Trade Date')

            # No sell or liquidate transactions in current data structure
            # Set Remaining to Sell to 0 since we only have buy transactions

            consolidated_positions.append(row)

    # Group 1: Received in Addepar == False (Top Level Owner + Instrument), Original Commitment = sum of buys
    build_positions(received_false, ['Top Level Owner', 'Instrument'])

    build_positions(received_true, ['Direct Owner Entity ID', 'Entity ID'])

    consolidated_df = pd.DataFrame(consolidated_positions)
    return consolidated_df


def merge_and_calculate_final_metrics(consolidated_investment_df, addepar_result, subscription_dates):
    """Merge data sources and calculate final financial metrics."""
    print("[STEP 5/7] Merging data sources and calculating metrics...")

    # Ensure Direct Owner Entity ID and Entity ID are numeric for consistent joins
    consolidated_investment_df = coerce_id_columns_to_int64(
        consolidated_investment_df)
    addepar_result = coerce_id_columns_to_int64(addepar_result)
    subscription_dates = coerce_id_columns_to_int64(subscription_dates)

    # Merge on normalized columns for case-insensitive join
    merged = consolidated_investment_df.merge(
        addepar_result,
        left_on=["Direct Owner Entity ID", "Entity ID"],
        right_on=["Direct Owner Entity ID", "Entity ID"],
        how="left"
    )

    # Add Subscription Approval Date
    # Handle subscription_dates merge carefully to preserve Top Level Owner column
    if not subscription_dates.empty:
        # For records with IDs, merge on IDs
        id_based_subscriptions = subscription_dates[subscription_dates['Direct Owner Entity ID'].notna(
        )]
        if not id_based_subscriptions.empty:
            merged = merged.merge(
                id_based_subscriptions[[
                    'Direct Owner Entity ID', 'Entity ID', 'Subscription Approval Date']],
                on=['Direct Owner Entity ID', 'Entity ID'],
                how='left'
            )

        # For records without IDs, merge on Top Level Owner and Instrument
        client_based_subscriptions = subscription_dates[subscription_dates['Direct Owner Entity ID'].isna(
        )]
        if not client_based_subscriptions.empty:
            # Create a temporary merge key for client-based records
            merged['temp_client_key'] = merged['Top Level Owner'].astype(
                str) + '|' + merged['Instrument'].astype(str)
            client_based_subscriptions['temp_client_key'] = client_based_subscriptions['Top Level Owner'].astype(
                str) + '|' + client_based_subscriptions['Investment'].astype(str)

            # Merge client-based subscriptions
            client_merge = merged.merge(
                client_based_subscriptions[[
                    'temp_client_key', 'Subscription Approval Date']],
                on='temp_client_key',
                how='left',
                suffixes=('', '_client')
            )

            # Update Subscription Approval Date where it's missing from ID-based merge
            mask_missing = merged['Subscription Approval Date'].isna()
            merged.loc[mask_missing, 'Subscription Approval Date'] = client_merge.loc[mask_missing,
                                                                                      'Subscription Approval Date_client']

            # Clean up temporary column
            merged = merged.drop(columns=['temp_client_key'])
    else:
        # If subscription_dates is empty, add empty column
        merged['Subscription Approval Date'] = pd.NA


    # import CA,TI sheet from Alts Info.xlsx
    ca_ti_df = pd.read_excel(
        "Alts Info.xlsx", sheet_name="CA,TI", engine="openpyxl")
    print(f"  [OK] Loaded {len(ca_ti_df)} CA,TI records")

    # for each row in merged, add to 'Total Contributions' the "Capital Calls prior to Transfer In" column value for the corresponding Direct Owner Entity ID and Entity ID
    # Efficiently add 'Capital Calls prior to Transfer In' and 'Capital Returned prior to Transfer IN'
    # to merged, acknowledging that not all items in merged will be in ca_ti_df.

    # Prepare ca_ti_df with IDs as keys for a left merge
    ca_ti_cols = [
        'Direct Owner Entity ID',
        'Entity ID',
        'Capital Calls prior to Transfer In',
        'Capital Returned prior to Transfer IN',
    ]
    ca_ti_short = ca_ti_df[ca_ti_cols].copy()

    # Merge CA,TI columns in as new columns (will be NaN if no match)
    merged = merged.merge(
        ca_ti_short,
        on=['Direct Owner Entity ID', 'Entity ID'],
        how='left',
        suffixes=('', '_ca_ti')
    )
    # Fill NaN with 0 for additive logic
    merged['Capital Calls prior to Transfer In'] = merged['Capital Calls prior to Transfer In'].fillna(0)
    merged['Capital Returned prior to Transfer IN'] = merged['Capital Returned prior to Transfer IN'].fillna(0)

    # Do the addition, preserving original NA if present (should be numeric, but robust to non-numeric)
    merged['Total Contributions'] = merged['Total Contributions'].fillna(0) + merged['Capital Calls prior to Transfer In']
    merged['Capital Returned'] = merged['Capital Returned'].fillna(0) + merged['Capital Returned prior to Transfer IN']

    # Remove the helper columns
    merged = merged.drop(columns=['Capital Calls prior to Transfer In', 'Capital Returned prior to Transfer IN'])

    # Calculate unfunded capital
    merged['Unfunded Capital'] = merged['Original Commitment'] - \
        merged['Total Contributions'] + merged['Capital Returned']
    
    print(f"  [OK] Merge complete - {len(merged)} total records\n")

    return merged


def format_and_save_excel(merged_data, investment_status_df, output_filename):
    """Format dates and save the final Excel file with proper formatting."""
    print("[STEP 6/7] Formatting data for Excel output...")

    # For positions without IDs (not in Addepar), use Investment as Instrument if Instrument is NaN
    no_ids_mask = merged_data['Direct Owner Entity ID'].isna(
    ) & merged_data['Entity ID'].isna()
    instrument_missing = merged_data['Instrument'].isna() | (merged_data['Instrument'].astype(
        str).str.strip() == '') | (merged_data['Instrument'].astype(str) == 'nan')
    if 'Investment' in merged_data.columns:
        merged_data.loc[no_ids_mask & instrument_missing,
                        'Instrument'] = merged_data.loc[no_ids_mask & instrument_missing, 'Investment']

    # Fill alts list fields with 0 for positions without IDs
    alts_list_cols = ['Total Commitments', 'Unfunded Capital',
                      'Total Contributions', 'Capital Returned', 'Mkt Value (USD)']
    for col in alts_list_cols:
        if col in merged_data.columns:
            merged_data.loc[no_ids_mask,
                            col] = merged_data.loc[no_ids_mask, col].fillna(0)

    # Select and reorder columns (removed Transaction Type since we now have one row per position)
    # Build column list conditionally to avoid missing-column errors

    # Ensure Legal Entity column exists - create it if it doesn't exist
    if 'Legal Entity' not in merged_data.columns:
        # Try to use 'Client/Top Level Owner' if it exists, otherwise use 'Top Level Owner'
        if 'Top Level Owner' in merged_data.columns:
            merged_data['Legal Entity'] = merged_data['Top Level Owner']
        else:
            merged_data['Legal Entity'] = merged_data['Top Level Owner']
    
    # Ensure Advisor column exists - create it as empty if it doesn't exist
    if 'Advisor' not in merged_data.columns:
        merged_data['Advisor'] = ''
    
    # Ensure Top Level Owner ID column exists - create it as empty if it doesn't exist
    if 'Top Level Owner ID' not in merged_data.columns:
        merged_data['Top Level Owner ID'] = ''

    base_cols = ["Top Level Owner", "Top Level Owner ID", "Legal Entity", "Advisor", "Account", "Instrument",  "Has all transactions", "Instruction Date", "Last Trade Date", "Subscription Approval Date", "Original Commitment",
                 'Total Contributions',
                 'Unfunded Capital',
                 'Capital Returned', 'Mkt Value (USD)', "Last Buy/Contribution", "Last Sell/Distribution", "is_open"]
    output_cols = base_cols.copy()
    if 'Fully Funded?' in merged_data.columns:
        output_cols.insert(-1, 'Fully Funded?')
    final_df = merged_data[output_cols].copy()  # Use .copy() to avoid SettingWithCopyWarning

    # Prepare Open Reason column (populated below for Open rows only)
    final_df['Open Reason'] = ''

    # Ensure Account column has empty string instead of NaN, 'nan', 'None', or None
    final_df['Account'] = final_df['Account'].astype(str).replace(
        ['nan', 'None'], '').replace({pd.NA: '', None: ''}).fillna('').str.strip()

    # Sort by Instruction Date (most recent last) and compute last dates per Account+Instrument
    final_df['Instruction Date'] = pd.to_datetime(
        final_df['Instruction Date'], errors='coerce')
    # Unify Last Trade Date as the max of per-row last trade, last buy/contribution, last sell/distribution
    per_row_last_trade = pd.to_datetime(
        final_df['Last Trade Date'], errors='coerce')
    last_buy = pd.to_datetime(final_df['Last Buy/Contribution'],
                              errors='coerce') if 'Last Buy/Contribution' in final_df.columns else pd.NaT
    last_sell = pd.to_datetime(final_df['Last Sell/Distribution'],
                               errors='coerce') if 'Last Sell/Distribution' in final_df.columns else pd.NaT
    final_df['Last Trade Date'] = per_row_last_trade
    if 'Last Buy/Contribution' in final_df.columns:
        final_df['Last Trade Date'] = final_df[[
            'Last Trade Date', 'Last Buy/Contribution']].max(axis=1)
    if 'Last Sell/Distribution' in final_df.columns:
        final_df['Last Trade Date'] = final_df[[
            'Last Trade Date', 'Last Sell/Distribution']].max(axis=1)
    # Compute last dates for each position
    final_df['Last Instruction Date'] = final_df.groupby(['Account', 'Instrument'])[
        'Instruction Date'].transform('max')
    final_df['Last Trade Date'] = pd.to_datetime(
        final_df['Last Trade Date'], errors='coerce')
    final_df['Last Trade Date'] = final_df.groupby(['Account', 'Instrument'])[
        'Last Trade Date'].transform('max')
    # Replace 'Instruction Date' column with 'Last Instruction Date' at the same position in the output
    cols_order = list(final_df.columns)
    if 'Instruction Date' in cols_order and 'Last Instruction Date' in cols_order:
        insert_idx = cols_order.index('Instruction Date')
        # Remove any existing occurrence to avoid duplicates
        cols_order = [c for c in cols_order if c not in [
            'Instruction Date', 'Last Instruction Date']]
        cols_order.insert(insert_idx, 'Last Instruction Date')
        final_df = final_df[cols_order]
    # Sort by the computed Last Instruction Date
    sort_col = 'Last Instruction Date' if 'Last Instruction Date' in final_df.columns else 'Instruction Date'
    final_df = final_df.sort_values(
        sort_col, ascending=True, na_position='first')
    if 'Has all transactions' in final_df.columns:
        # if its not empty and greater than 0, map to true, otherwise map to false
        final_df['Has all transactions'] = final_df['Has all transactions'].notna() & (
            final_df['Has all transactions'] > 0)
        final_df['Has all transactions'] = final_df['Has all transactions'].map(
            {True: 'No', False: 'Yes'})
        final_df = final_df.sort_values(
            ['Has all transactions', sort_col], ascending=[False, True], na_position='first')
    date_columns = ["Instruction Date", "Last Instruction Date", "Last Trade Date",
                    "Subscription Approval Date", "Last Buy/Contribution", "Last Sell/Distribution"]
    for col in date_columns:
        if col in final_df.columns:
            final_df[col] = pd.to_datetime(
                final_df[col], errors='coerce').dt.strftime('%m/%d/%Y')
            final_df[col] = final_df[col].fillna("")

    # Delete any previous 'Alts Log *.xlsx' files in the root directory
    for old_file in glob.glob('Alts Log *.xlsx'):
        try:
            os.remove(old_file)
        except Exception as e:
            print(f"Could not delete {old_file}: {e}")

    # Propagate is_open to all rows for each (Account, Instrument)
    final_df['is_open'] = final_df.groupby(['Top Level Owner', 'Instrument'])[
        'is_open'].transform('max')

    # Convert Unfunded Capital to numeric for comparison

    print("[STEP 6/7] Calculating open reasons for positions...")
    # Check for incomplete transactions in investment_status_df first (higher priority than Unfunded Capital)
    for idx, row in final_df.iterrows():
        if row['is_open'] != False:
            # Find matching rows in investment_status_df
            if row['Account'] and row['Instrument']:
                # For positions with Account and Instrument, compare using Instrument_Match for consistency
                matching_rows = investment_status_df[
                    (investment_status_df['Account'].astype(str).str.strip().str.lower() == str(row['Account']).strip().lower()) &
                    (investment_status_df['Instrument_Match'].astype(
                        str).str.strip().str.lower() == str(row['Instrument']).strip().lower())
                ]
            else:
                # For positions without Account or without Addepar IDs, use Top Level Owner and Instrument
                matching_rows = investment_status_df[
                    (investment_status_df['Top Level Owner'].astype(str).str.strip().str.lower() == str(row['Top Level Owner']).strip().lower()) &
                    (investment_status_df['Instrument_Match'].astype(
                        str).str.strip().str.lower() == str(row['Instrument']).strip().lower())
                ]
            # Determine Open Reason per new rules
            # - if not "Initial Funding Received": in buy process
            # - if "Initial Funding Received" and not "Fully Funded?": unfunded capital

            reason_set = False
            # Check if Initial Trade Executed is False (0 or 'false')
            not_initial_received = matching_rows['Initial Trade Executed'].apply(is_false_value)
            if not_initial_received.any():
                final_df.loc[idx, 'Open Reason'] = 'In Buy Process'
            # Check unfunded if not already set
            else:
                addepar_received_col = 'Received in Addepar'
                # Check if Received in Addepar is False (0 or 'false')
                not_addepar_received = matching_rows[addepar_received_col].apply(is_false_value)
                if not_addepar_received.any():
                    final_df.loc[idx,
                                 'Open Reason'] = 'Not Received in Addepar'
                else:
                    funded_col = 'Fully Funded?'
                    # Check if Fully Funded is False (0 or 'false')
                    not_fully_funded = matching_rows[funded_col].apply(is_false_value)
                    if not_fully_funded.any():
                        final_df.loc[idx, 'Open Reason'] = 'Unfunded Capital'

    # Split into open and close sheets based on is_open
    print("[STEP 6/7] Splitting into Open and Closed sheets...")
    open_df = final_df[final_df['is_open'] == True].copy()
    open_df = open_df.sort_values(['Top Level Owner', 'Account', 'Instrument'])

    close_df = final_df[final_df['is_open'] == False].copy()
    close_df = close_df.sort_values(
        ['Top Level Owner', 'Account', 'Instrument'])
    
    print(f"  [OK] Open positions: {len(open_df)}")
    print(f"  [OK] Closed positions: {len(close_df)}")

    # Drop funding-status column and is_open from output (keep Account in both sheets)
    cols_to_drop = [col for col in ['Completed?',
                                    'Fully Funded?', 'is_open'] if col in open_df.columns]
    open_df = open_df.drop(columns=cols_to_drop)
    close_df = close_df.drop(columns=cols_to_drop)

    # Do not include Open Reason in the Closed sheet
    if 'Open Reason' in close_df.columns:
        close_df = close_df.drop(columns=['Open Reason'])

    # Ensure column naming is Top Level Owner in both sheets and reorder Open

    # Reorder columns in Open sheet: place Open Reason after Top Level Owner, Top Level Owner ID, Legal Entity, Advisor, Account, Instrument
    prefix = [col for col in ['Top Level Owner', 'Top Level Owner ID', 'Legal Entity', 'Advisor', 'Account',
                              'Instrument', 'Open Reason'] if col in open_df.columns]
    others = [c for c in open_df.columns if c not in prefix]
    open_df = open_df[prefix + others]

    def _format_sheet(writer, sheet_name, sheet_df):
        worksheet = writer.sheets[sheet_name]
        worksheet.sheet_view.zoomScale = 100
        # sort by Top Level Owner, Account, Instrument
        all_currency_cols = ["Original Commitment", 'Mkt Value (USD)',
                             'Total Commitments', 'Unfunded Capital',
                             'Total Contributions',
                             'Capital Returned']
        currency_cols = [
            col for col in all_currency_cols if col in sheet_df.columns]

        # if colname contains "Date", convert to date
        for col_idx, col in enumerate(sheet_df.columns, 1):
            if "Date" in col:
                sheet_df[col] = pd.to_datetime(sheet_df[col], errors='coerce')
                for cell in worksheet[chr(64+col_idx)][1:]:
                    cell.number_format = 'm/d/yyyy'
        # Format currency columns
        for col_idx, col in enumerate(sheet_df.columns, 1):
            if col in currency_cols:
                for cell in worksheet[chr(64+col_idx)][1:]:  # skip header
                    cell.number_format = '"$"#,##0'
        # Set column widths based on formatted values
        for col_idx, col in enumerate(sheet_df.columns, 1):
            worksheet.column_dimensions[chr(64+col_idx)].width = 25

        # Add conditional formatting for Open Reason column (only for Open sheet)
        if sheet_name == "Open" and 'Open Reason' in sheet_df.columns:
            open_reason_col_idx = None
            for col_idx, col in enumerate(sheet_df.columns, 1):
                if col == 'Open Reason':
                    open_reason_col_idx = col_idx
                    break

            if open_reason_col_idx:
                from openpyxl.styles import PatternFill

                # Define colors
                light_orange = PatternFill(
                    start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                light_green = PatternFill(
                    start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                light_red = PatternFill(
                    start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                # Apply conditional formatting to Open Reason cells
                for row in range(2, worksheet.max_row + 1):  # Skip header row
                    cell = worksheet.cell(row=row, column=open_reason_col_idx)
                    if cell.value == "Unfunded Capital":
                        cell.fill = light_orange
                    elif cell.value == "In Buy Process":
                        cell.fill = light_green
                    elif cell.value == "Not Received in Addepar":
                        cell.fill = light_red

        # Add filters to all columns
        worksheet.auto_filter.ref = worksheet.dimensions

        # Freeze first row (header) and first column
        worksheet.freeze_panes = 'B2'

    print("\n[STEP 7/7] Saving Excel file...")
    # Save to Excel with adjusted column widths and currency formatting
    with pd.ExcelWriter(f"{output_filename}", engine="openpyxl") as writer:
        print(f"  - Writing Open sheet ({len(open_df)} rows)...")
        open_df.to_excel(writer, index=False, sheet_name="Open")
        print(f"  - Writing Closed sheet ({len(close_df)} rows)...")
        close_df.to_excel(writer, index=False, sheet_name="Closed")
        print("  - Formatting Open sheet...")
        _format_sheet(writer, "Open", open_df)
        print("  - Formatting Closed sheet...")
        _format_sheet(writer, "Closed", close_df)
    
    print(f"  [OK] File saved: {output_filename}\n")
    print("=" * 50)
    print("[SUCCESS] Process completed successfully!")
    print("=" * 50)
    os.startfile(f"{output_filename}")


def main():
    """Main application function."""
    print("\n" + "=" * 50)
    print("ALTS LOG GENERATION PROCESS")
    print("=" * 50 + "\n")
    
    # Determine today's date for filename
    output_filename = f'Alts Log {datetime.now().strftime("%m-%d-%Y")}.xlsx'

    # If today's log already exists, open and exit
    if os.path.exists(output_filename):
        print(f"[OK] Today's log already exists: {output_filename}")
        print("  Opening existing file...\n")
        os.startfile(output_filename)
        exit()

    print("Starting new log generation process...\n")
    
    # Set up environment and API credentials
    print("[SETUP] Configuring environment and API credentials...")
    api_key, api_secret, firm_id, base_url, start_date, end_date = set_up_environment()
    print(f"  [OK] Environment configured (Period: {start_date} to {end_date})\n")
    
    transactions_df, addepar_result = fetch_and_process_addepar_data(
        api_key, api_secret, firm_id, base_url, start_date, end_date
    )
    alts_list_df = fetch_and_process_alts_list_data(
        api_key, api_secret, firm_id, base_url, start_date, end_date
    )
    if transactions_df is None or alts_list_df is None:
        print("\n[ERROR] Failed to fetch required data. Please check your API connection and credentials.")
        input("\nPress Enter to exit...")
        return

    investment_status_df, subscription_dates, complete_summary = process_alts_info_data()
    
    print("[STEP 4/7] Merging alts list with investment status...")
    # merge alts_list_df and investment_status_df on Direct Owner Entity ID and Entity ID
    consolidated_investment_df = investment_status_df.merge(
        alts_list_df, on=["Direct Owner Entity ID", "Entity ID"], how="left")
    print(f"  [OK] Merged {len(consolidated_investment_df)} records\n")

    merged_data = merge_and_calculate_final_metrics(
        consolidated_investment_df, addepar_result, subscription_dates)
    merged_data = merged_data.merge(
        complete_summary, on=["Direct Owner Entity ID", "Entity ID"], how="left")
    format_and_save_excel(merged_data, investment_status_df, output_filename)


if __name__ == "__main__":
    main()
